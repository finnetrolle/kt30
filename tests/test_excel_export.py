import unittest
from io import BytesIO
from importlib.util import module_from_spec, spec_from_file_location
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

EXCEL_EXPORT_PATH = Path(__file__).resolve().parents[1] / "excel_export.py"
EXCEL_EXPORT_SPEC = spec_from_file_location("excel_export_real", EXCEL_EXPORT_PATH)
excel_export_real = module_from_spec(EXCEL_EXPORT_SPEC)
assert EXCEL_EXPORT_SPEC.loader is not None
EXCEL_EXPORT_SPEC.loader.exec_module(excel_export_real)

calculate_project_duration_with_parallel = excel_export_real.calculate_project_duration_with_parallel
create_wbs_excel = excel_export_real.create_wbs_excel
normalize_role = excel_export_real.normalize_role


class ExcelExportTests(unittest.TestCase):
    def setUp(self):
        self.sample_result = {
            "analysis": {
                "functional_requirements": [
                    {
                        "id": "FR-1",
                        "name": "Сбор и обработка данных",
                        "description": "Система должна собирать и обрабатывать данные из внешних источников.",
                    }
                ]
            },
            "wbs": {
                "phases": [
                    {
                        "id": "1",
                        "name": "Фаза",
                        "estimated_hours": 16,
                        "skills_required": ["Архитектор"],
                        "work_packages": [
                            {
                                "id": "1.1",
                                "name": "Пакет",
                                "estimated_hours": 16,
                                "duration_days": 1,
                                "skills_required": ["Проектный менеджер"],
                                "requirement_ids": ["FR-1"],
                                "tasks": [
                                    {
                                        "id": "1.1.1",
                                        "name": "Бэкенд",
                                        "estimated_hours": 8,
                                        "duration_days": 1,
                                        "skills_required": ["Backend-разработчик"],
                                    },
                                    {
                                        "id": "1.1.2",
                                        "name": "Тест",
                                        "estimated_hours": 8,
                                        "duration_days": 2,
                                        "skills_required": ["QA-инженер"],
                                    },
                                ],
                            }
                        ],
                    }
                ]
            }
        }

    def _load_workbook(self) -> tuple:
        excel_bytes = create_wbs_excel(self.sample_result)
        workbook = load_workbook(filename=BytesIO(excel_bytes.getvalue()), data_only=False)
        return workbook, workbook["WBS"], workbook["Ставки профессий"]

    def test_wbs_rolls_up_parents_and_totals_only_sum_leaf_rows(self):
        workbook, ws_wbs, ws_rates = self._load_workbook()
        self.assertEqual(workbook.sheetnames, ["WBS", "Ставки профессий"])

        header_to_col = {ws_wbs.cell(1, col).value: col for col in range(1, ws_wbs.max_column + 1)}
        quote_col = header_to_col["Цитата из требований"]
        backend_col = header_to_col["Backend-разработчик"]
        qa_col = header_to_col["QA-инженер"]
        pm_col = header_to_col["Проектный менеджер"]
        cost_col = header_to_col["Стоимость (руб)"]
        backend_letter = get_column_letter(backend_col)
        qa_letter = get_column_letter(qa_col)
        pm_letter = get_column_letter(pm_col)
        cost_letter = get_column_letter(cost_col)

        self.assertEqual(ws_wbs.freeze_panes, "E2")
        self.assertIsNone(ws_wbs.cell(7, 1).value)
        self.assertEqual(ws_wbs.cell(7, 2).value, "ИТОГО")
        self.assertIn("FR-1: Сбор и обработка данных.", ws_wbs.cell(3, quote_col).value)
        self.assertIn("FR-1: Сбор и обработка данных.", ws_wbs.cell(4, quote_col).value)

        self.assertEqual(ws_wbs.cell(2, backend_col).value, f"=SUM({backend_letter}3)")
        self.assertEqual(ws_wbs.cell(3, backend_col).value, f"=SUM({backend_letter}4:{backend_letter}5)")
        self.assertEqual(ws_wbs.cell(3, qa_col).value, f"=SUM({qa_letter}4:{qa_letter}5)")
        self.assertEqual(ws_wbs.cell(3, pm_col).value, f"=SUM({pm_letter}4:{pm_letter}5)")
        self.assertEqual(ws_wbs.cell(7, backend_col).value, f"=SUM({backend_letter}4:{backend_letter}5)")
        self.assertEqual(ws_wbs.cell(7, qa_col).value, f"=SUM({qa_letter}4:{qa_letter}5)")
        self.assertEqual(ws_wbs.cell(7, cost_col).value, f"=SUM({cost_letter}4:{cost_letter}5)")

        rate_roles = [ws_rates.cell(row, 1).value for row in range(2, ws_rates.max_row) if ws_rates.cell(row, 1).value]
        self.assertEqual(set(rate_roles), {"Backend-разработчик", "Проектный менеджер", "QA-инженер"})
        self.assertEqual(len(rate_roles), 3)

    def test_export_does_not_create_gantt_sheet(self):
        workbook, _ws_wbs, _ws_rates = self._load_workbook()
        self.assertNotIn("Диаграмма Гантта", workbook.sheetnames)

    def test_duration_calculation_uses_child_span(self):
        duration = calculate_project_duration_with_parallel(self.sample_result["wbs"])
        self.assertEqual(duration["total_days"], 3)
        self.assertEqual(duration["total_weeks"], 1)
        self.assertEqual(duration["phase_durations"], {"1": 3})

    def test_role_normalization_collapses_technology_specific_variants(self):
        self.assertEqual(normalize_role("PostgreSQL Engineer"), "Администратор БД")
        self.assertEqual(normalize_role("SQL Developer"), "Администратор БД")
        self.assertEqual(normalize_role("DBA"), "Администратор БД")
        self.assertEqual(normalize_role("Архитектор безопасности"), "Архитектор")
        self.assertEqual(normalize_role("Pentester"), "Специалист ИБ")
        self.assertEqual(normalize_role("Performance Engineer"), "QA-инженер")
        self.assertEqual(normalize_role("Интеграционный разработчик"), "Backend-разработчик")


if __name__ == "__main__":
    unittest.main()
