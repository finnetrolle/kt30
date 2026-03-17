"""
Excel export module for WBS (Work Breakdown Structure).
Generates Excel files with work packages, roles, hours, costs, and Gantt chart.
"""
import logging
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from wbs_utils import canonicalize_wbs_result

logger = logging.getLogger(__name__)

# Default role rates (rubles per hour)
# These can be customized on the second sheet of the Excel file
DEFAULT_ROLE_RATES = {
    "Проектный менеджер": 1500,
    "Системный аналитик": 1300,
    "Архитектор": 2000,
    "Frontend-разработчик": 1000,
    "Backend-разработчик": 1100,
    "Data Engineer": 1200,
    "QA-инженер": 900,
    "DevOps": 1500,
    "Администратор БД": 1400,
    "Специалист ИБ": 1300,
    "Технический писатель": 700,
    "Дизайнер": 1000,
}

# Role aliases for normalization
ROLE_ALIASES = {
    "Frontend разработчик": "Frontend-разработчик",
    "Разработчик (Frontend)": "Frontend-разработчик",
    "Backend разработчик": "Backend-разработчик",
    "Разработчик (Backend)": "Backend-разработчик",
    "Full-stack разработчик": "Backend-разработчик",
    "Разработчик (Full-stack)": "Backend-разработчик",
    "QA инженер": "QA-инженер",
    "Тестировщик": "QA-инженер",
    "Тестировщик (QA)": "QA-инженер",
    "QA": "QA-инженер",
    "PM": "Проектный менеджер",
    "Руководитель проекта": "Проектный менеджер",
    "Scrum мастер": "Проектный менеджер",
    "БА": "Системный аналитик",
    "СА": "Системный аналитик",
    "БД": "Администратор БД",
    "DBA": "Администратор БД",
    "DevOps инженер": "DevOps",
    "BI-разработчик": "Data Engineer",
    "Инженер БД": "Администратор БД",
    "Архитектор ПО": "Архитектор",
    "Архитектор данных": "Архитектор",
    "Архитектор безопасности": "Архитектор",
    "Интеграционный архитектор": "Архитектор",
    "Интеграционный разработчик": "Backend-разработчик",
    "Инженер инфраструктуры": "DevOps",
    "Инженер надежности": "DevOps",
    "Инженер развертывания": "DevOps",
    "Performance Engineer": "QA-инженер",
    "Инженер по нагрузке": "QA-инженер",
    "Pentester": "Специалист ИБ",
}


def calculate_project_duration_with_parallel(wbs_data: dict) -> dict:
    """Calculate actual project duration considering parallel execution.
    
    Args:
        wbs_data: WBS data dictionary
        
    Returns:
        Dictionary with duration information:
        - total_days: total working days considering parallel execution
        - total_weeks: total weeks
        - phase_durations: dict of phase_id -> actual days
    """
    if not wbs_data or 'phases' not in wbs_data:
        return {'total_days': 0, 'total_weeks': 0, 'phase_durations': {}}
    
    item_end_days = {}
    phase_durations = {}
    current_day = 0
    
    for phase in wbs_data.get('phases', []):
        phase_id = phase.get('id', '')
        phase_duration = phase.get('duration', 0) or phase.get('duration_days', 0)
        if not phase_duration:
            hours = phase.get('estimated_hours', 0)
            phase_duration = max(1, hours // 8)
        phase_start = current_day
        
        # Track parallel work packages
        parallel_wp_end = phase_start
        sequential_wp_start = phase_start
        
        for wp in phase.get('work_packages', []):
            wp_id = wp.get('id', '')
            wp_duration = wp.get('duration_days', 0)
            if not wp_duration:
                hours = wp.get('estimated_hours', 0)
                wp_duration = max(1, hours // 8)
            
            wp_dependencies = wp.get('dependencies', [])
            can_parallel = wp.get('can_start_parallel', False)
            
            dep_end_day = 0
            for dep_id in wp_dependencies:
                if dep_id in item_end_days:
                    dep_end_day = max(dep_end_day, item_end_days[dep_id])
            
            if can_parallel:
                wp_start = dep_end_day if dep_end_day > 0 else phase_start
            else:
                wp_start = max(sequential_wp_start, dep_end_day)
            
            # Process tasks
            parallel_task_end = wp_start
            sequential_task_start = wp_start
            
            for task in wp.get('tasks', []):
                task_id = task.get('id', '')
                task_duration = task.get('duration_days', 0)
                if not task_duration:
                    hours = task.get('estimated_hours', 0)
                    task_duration = max(1, hours // 8)
                
                task_dependencies = task.get('dependencies', [])
                task_can_parallel = task.get('can_start_parallel', False)
                
                task_dep_end = 0
                for dep_id in task_dependencies:
                    if dep_id in item_end_days:
                        task_dep_end = max(task_dep_end, item_end_days[dep_id])
                
                if task_can_parallel:
                    task_start = task_dep_end if task_dep_end > 0 else wp_start
                else:
                    task_start = max(sequential_task_start, task_dep_end)
                
                task_end = task_start + task_duration
                item_end_days[task_id] = task_end
                
                if not task_can_parallel:
                    sequential_task_start = task_end
                
                parallel_task_end = max(parallel_task_end, task_end)
            
            wp_end = max(wp_start + wp_duration, parallel_task_end)
            item_end_days[wp_id] = wp_end
            
            if not can_parallel:
                sequential_wp_start = wp_end
            
            parallel_wp_end = max(parallel_wp_end, wp_end)
        
        actual_phase_end = parallel_wp_end if phase.get('work_packages') else phase_start + phase_duration
        phase_durations[phase_id] = actual_phase_end - phase_start
        current_day = actual_phase_end
    
    total_days = current_day
    total_weeks = max(1, (total_days + 4) // 5)  # Ceiling division for weeks
    
    return {
        'total_days': total_days,
        'total_weeks': total_weeks,
        'phase_durations': phase_durations
    }


def build_dependencies_matrix(wbs_data: dict) -> list:
    """Build a dependencies matrix showing what each task depends on and what can run in parallel.
    
    Args:
        wbs_data: WBS data dictionary
        
    Returns:
        List of dependency entries with task_id, depends_on, and parallel_with
    """
    matrix = []
    
    if not wbs_data or 'phases' not in wbs_data:
        return matrix
    
    # Collect all tasks and work packages with their parallel info
    all_items = {}
    parallel_groups = {}  # Group items that can run in parallel
    
    for phase in wbs_data.get('phases', []):
        phase_id = phase.get('id', '')
        
        for wp in phase.get('work_packages', []):
            wp_id = wp.get('id', '')
            wp_can_parallel = wp.get('can_start_parallel', False)
            wp_dependencies = wp.get('dependencies', [])
            
            all_items[wp_id] = {
                'id': wp_id,
                'name': wp.get('name', ''),
                'type': 'work_package',
                'can_parallel': wp_can_parallel,
                'dependencies': wp_dependencies,
                'phase_id': phase_id
            }
            
            # Group parallel work packages by phase
            if wp_can_parallel:
                if phase_id not in parallel_groups:
                    parallel_groups[phase_id] = []
                parallel_groups[phase_id].append(wp_id)
            
            for task in wp.get('tasks', []):
                task_id = task.get('id', '')
                task_can_parallel = task.get('can_start_parallel', False)
                task_dependencies = task.get('dependencies', [])
                
                all_items[task_id] = {
                    'id': task_id,
                    'name': task.get('name', ''),
                    'type': 'task',
                    'can_parallel': task_can_parallel,
                    'dependencies': task_dependencies,
                    'wp_id': wp_id,
                    'phase_id': phase_id
                }
                
                # Group parallel tasks by work package
                if task_can_parallel:
                    if wp_id not in parallel_groups:
                        parallel_groups[wp_id] = []
                    parallel_groups[wp_id].append(task_id)
    
    # Build the matrix
    for item_id, item in all_items.items():
        # Find what can run in parallel with this item
        parallel_with = []
        
        if item['can_parallel']:
            # Find the group key (phase for work packages, wp for tasks)
            group_key = item.get('phase_id') if item['type'] == 'work_package' else item.get('wp_id')
            
            if group_key in parallel_groups:
                parallel_with = [pid for pid in parallel_groups[group_key] if pid != item_id]
        
        matrix.append({
            'task_id': item_id,
            'task_name': item['name'],
            'type': item['type'],
            'depends_on': item['dependencies'],
            'parallel_with': parallel_with
        })
    
    return matrix


def normalize_role(role_name: str) -> str:
    """Normalize role name using aliases.
    
    Args:
        role_name: Original role name
        
    Returns:
        Normalized role name
    """
    if not role_name:
        return role_name

    role_name = str(role_name).strip()

    # Check aliases first
    if role_name in ROLE_ALIASES:
        return ROLE_ALIASES[role_name]

    normalized_key = re.sub(r"[^a-zA-Zа-яА-Я0-9+]+", " ", role_name.casefold()).strip()

    keyword_groups = [
        (
            "Проектный менеджер",
            ("project manager", "project lead", "product manager", "delivery manager", "pm", "scrum", "менеджер проекта", "менеджер продукта", "руководитель проекта", "руководитель программы", "проектн"),
        ),
        (
            "Системный аналитик",
            ("business analyst", "system analyst", "systems analyst", "analyst", "product analyst", "аналитик", "бизнес аналитик", "системный аналитик", "бизнес аналитик", "бизнес-аналитик"),
        ),
        (
            "Архитектор",
            ("architect", "solution architect", "software architect", "enterprise architect", "data architect", "integration architect", "архитектор", "solution architecture"),
        ),
        (
            "Специалист ИБ",
            ("security", "infosec", "cyber", "pentest", "penetration", "appsec", "devsecops", "secops", "иб", "информационной безопасности", "кибербезопас", "пентест", "безопасност"),
        ),
        (
            "Администратор БД",
            ("postgres", "postgresql", "sql", "oracle", "mysql", "clickhouse", "greenplum", "mongodb", "database", "database administrator", "database engineer", "dba", "db admin", "db engineer", "субд", "бд", "база данных"),
        ),
        (
            "DevOps",
            ("devops", "sre", "platform engineer", "infra", "infrastructure", "site reliability", "release engineer", "deployment", "kubernetes", "docker", "terraform", "ansible", "ci cd", "ci/cd", "инфраструкт", "надежност", "развертыван", "эксплуатац"),
        ),
        (
            "Data Engineer",
            ("data engineer", "analytics engineer", "etl", "elt", "dwh", "warehouse", "bi", "spark", "hadoop", "airflow", "data platform", "данных", "витрин", "хранилищ", "аналитическ"),
        ),
        (
            "Frontend-разработчик",
            ("frontend", "front end", "react", "vue", "angular", "ui", "ux", "web", "веб", "фронтенд", "интерфейс"),
        ),
        (
            "Backend-разработчик",
            ("backend", "back end", "full stack", "fullstack", "developer", "engineer", "api", "service", "integration developer", "бэкенд", "сервер", "интеграционн", "разработчик", "программист"),
        ),
        (
            "QA-инженер",
            ("qa", "quality assurance", "test engineer", "tester", "testing", "performance engineer", "load engineer", "тест", "qa инженер", "контроль качества", "нагруз"),
        ),
        (
            "Технический писатель",
            ("technical writer", "documentation", "doc writer", "writer", "документац", "технический писатель"),
        ),
        (
            "Дизайнер",
            ("designer", "product designer", "ui designer", "ux designer", "дизайнер"),
        ),
    ]

    for canonical_role, keywords in keyword_groups:
        if any(keyword in normalized_key for keyword in keywords):
            return canonical_role

    # Check if role exists in default rates
    if role_name in DEFAULT_ROLE_RATES:
        return role_name

    # Return original if no normalization found
    return role_name


def get_role_rate(role_name: str, custom_rates: dict = None) -> int:
    """Get hourly rate for a role.
    
    Args:
        role_name: Name of the role
        custom_rates: Optional dictionary of custom rates
        
    Returns:
        Hourly rate for the role
    """
    normalized = normalize_role(role_name)

    if custom_rates:
        if role_name in custom_rates:
            return custom_rates[role_name]
        if normalized in custom_rates:
            return custom_rates[normalized]

        normalized_custom_rates = {normalize_role(key): value for key, value in custom_rates.items()}
        if normalized in normalized_custom_rates:
            return normalized_custom_rates[normalized]

    if normalized in DEFAULT_ROLE_RATES:
        return DEFAULT_ROLE_RATES[normalized]
    
    # Default rate for unknown roles
    return 1000


def extract_all_roles(wbs_data: dict) -> list:
    """Extract all unique roles from WBS data.
    
    Args:
        wbs_data: WBS data dictionary
        
    Returns:
        Sorted list of unique role names
    """
    roles = set()
    
    if not wbs_data or 'phases' not in wbs_data:
        return sorted(list(roles))
    
    for phase in wbs_data.get('phases', []):
        for wp in phase.get('work_packages', []):
            # Get roles from skills_required
            for skill in wp.get('skills_required', []):
                roles.add(normalize_role(skill))
            
            # Get roles from tasks
            for task in wp.get('tasks', []):
                for skill in task.get('skills_required', []):
                    roles.add(normalize_role(skill))
    
    return sorted(list(roles))


def extract_all_work_items(wbs_data: dict) -> list:
    """Extract all work items (phases, work packages, tasks) from WBS.
    
    Args:
        wbs_data: WBS data dictionary
        
    Returns:
        List of work items with their properties
    """
    items = []
    
    if not wbs_data or 'phases' not in wbs_data:
        return items
    
    for phase in wbs_data.get('phases', []):
        # Add phase as a work item
        items.append({
            'id': phase.get('id', ''),
            'name': phase.get('name', ''),
            'type': 'phase',
            'hours': phase.get('estimated_hours', 0),
            'skills': phase.get('skills_required', []),
            'description': phase.get('description', ''),
            'requirement_ids': [],
            'level': 0,
            'parent_id': None,
            'can_start_parallel': False,
            'dependencies': []
        })
        
        for wp in phase.get('work_packages', []):
            # Add work package
            items.append({
                'id': wp.get('id', ''),
                'name': wp.get('name', ''),
                'type': 'work_package',
                'hours': wp.get('estimated_hours', 0),
                'skills': wp.get('skills_required', []),
                'description': wp.get('description', ''),
                'requirement_ids': wp.get('requirement_ids', []),
                'level': 1,
                'parent_id': phase.get('id', ''),
                'can_start_parallel': wp.get('can_start_parallel', False),
                'dependencies': wp.get('dependencies', [])
            })
            
            # Add tasks
            for task in wp.get('tasks', []):
                items.append({
                    'id': task.get('id', ''),
                    'name': task.get('name', ''),
                    'type': 'task',
                    'hours': task.get('estimated_hours', 0),
                    'skills': task.get('skills_required', []) or wp.get('skills_required', []),
                    'description': task.get('description', ''),
                    'requirement_ids': task.get('requirement_ids', []) or wp.get('requirement_ids', []),
                    'level': 2,
                    'parent_id': wp.get('id', ''),
                    'can_start_parallel': task.get('can_start_parallel', False),
                    'dependencies': task.get('dependencies', [])
                })
    
    return items


def distribute_hours_by_role(hours: int, skills: list) -> dict:
    """Distribute hours among roles.
    
    Simple distribution: equal split among all required skills/roles.
    
    Args:
        hours: Total hours for the work item
        skills: List of required skills/roles
        
    Returns:
        Dictionary mapping roles to hours
    """
    if not skills or hours <= 0:
        return {}
    
    normalized_skills = [normalize_role(s) for s in skills]
    unique_skills = list(set(normalized_skills))
    
    if not unique_skills:
        return {}
    
    # Equal distribution
    hours_per_role = hours / len(unique_skills)
    
    return {role: hours_per_role for role in unique_skills}


def build_children_index(work_items: list) -> dict:
    """Build a parent -> children index for work items."""
    children_by_parent = {}

    for item in work_items:
        parent_id = item.get('parent_id')
        if not parent_id:
            continue
        children_by_parent.setdefault(parent_id, []).append(item)

    return children_by_parent


def build_sum_formula(column_letter: str, row_numbers: list) -> str:
    """Build a compact SUM formula for arbitrary row numbers."""
    if not row_numbers:
        return "=0"

    sorted_rows = sorted(set(row_numbers))
    ranges = []
    start = sorted_rows[0]
    end = start

    for row_number in sorted_rows[1:]:
        if row_number == end + 1:
            end = row_number
            continue
        if start == end:
            ranges.append(f"{column_letter}{start}")
        else:
            ranges.append(f"{column_letter}{start}:{column_letter}{end}")
        start = end = row_number

    if start == end:
        ranges.append(f"{column_letter}{start}")
    else:
        ranges.append(f"{column_letter}{start}:{column_letter}{end}")

    return "=SUM(" + ",".join(ranges) + ")"


def build_requirement_lookup(result_data: dict) -> dict:
    """Build a lookup for requirement text by requirement id."""
    lookup = {}
    analysis = result_data.get("analysis", {}) if isinstance(result_data, dict) else {}

    for collection_name in ("functional_requirements", "non_functional_requirements"):
        for requirement in analysis.get(collection_name, []) or []:
            requirement_id = str(requirement.get("id", "")).strip()
            if not requirement_id:
                continue
            name = str(requirement.get("name", "")).strip()
            description = str(requirement.get("description", "")).strip()

            if name and description and description != name:
                text = f"{requirement_id}: {name}. {description}"
            elif name:
                text = f"{requirement_id}: {name}"
            elif description:
                text = f"{requirement_id}: {description}"
            else:
                continue

            lookup[requirement_id] = text

    return lookup


def build_requirement_excerpt(item: dict, requirement_lookup: dict, children_by_parent: dict) -> str:
    """Build an excerpt showing which requirement(s) this work item implements."""
    requirement_ids = [str(req_id).strip() for req_id in item.get("requirement_ids", []) if str(req_id).strip()]

    if not requirement_ids and item.get("id") in children_by_parent:
        for child in children_by_parent.get(item["id"], []):
            for req_id in child.get("requirement_ids", []):
                req_id = str(req_id).strip()
                if req_id and req_id not in requirement_ids:
                    requirement_ids.append(req_id)

    excerpts = []
    for requirement_id in requirement_ids:
        text = requirement_lookup.get(requirement_id)
        if text and text not in excerpts:
            excerpts.append(text)

    if excerpts:
        if len(excerpts) > 3:
            excerpts = excerpts[:3] + [f"... еще {len(excerpts) - 3} треб."]
        return "\n".join(excerpts)

    fallback = str(item.get("description", "")).strip() or str(item.get("name", "")).strip()
    return fallback


def create_wbs_excel(result_data: dict, custom_rates: dict = None) -> BytesIO:
    """Create Excel file with WBS data.
    
    Args:
        result_data: Analysis result data
        custom_rates: Optional dictionary of custom role rates
        
    Returns:
        BytesIO object containing the Excel file
    """
    logger.info("Creating WBS Excel export...")
    
    wb = Workbook()
    
    # Get WBS data
    normalized_result = canonicalize_wbs_result(result_data)
    wbs_data = normalized_result.get('wbs', {}) if normalized_result else {}
    requirement_lookup = build_requirement_lookup(normalized_result)
    
    # Extract roles and work items
    all_roles = extract_all_roles(wbs_data)
    work_items = extract_all_work_items(wbs_data)
    
    # Use default rates if no custom rates provided
    rates = custom_rates or DEFAULT_ROLE_RATES.copy()
    
    # === Sheet 1: WBS with hours and costs ===
    ws_wbs = wb.active
    ws_wbs.title = "WBS"
    
    # Define styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    phase_font = Font(bold=True, size=11)
    phase_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    wp_font = Font(bold=True, size=10)
    wp_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    task_font = Font(size=10)
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create headers
    headers = ["№ работы", "Название работы", "Цитата из требований", "Паралл."] + all_roles + ["Стоимость (руб)"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_wbs.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Set column widths
    ws_wbs.column_dimensions['A'].width = 12  # ID
    ws_wbs.column_dimensions['B'].width = 40  # Name
    ws_wbs.column_dimensions['C'].width = 60  # Requirement quote
    ws_wbs.column_dimensions['D'].width = 8   # Parallel indicator
    
    for i, role in enumerate(all_roles, 5):
        col_letter = get_column_letter(i)
        ws_wbs.column_dimensions[col_letter].width = 15
    
    ws_wbs.column_dimensions[get_column_letter(len(headers))].width = 18  # Cost
    
    children_by_parent = build_children_index(work_items)
    parent_ids = set(children_by_parent)
    leaf_ids = {item['id'] for item in work_items if item['id'] not in parent_ids}
    row_by_item_id = {item['id']: index for index, item in enumerate(work_items, start=2)}

    # Add work items
    row = 2
    for item in work_items:
        # Determine style based on item type
        if item['type'] == 'phase':
            font = phase_font
            fill = phase_fill
        elif item['type'] == 'work_package':
            font = wp_font
            fill = wp_fill
        else:
            font = task_font
            fill = None
        
        # ID column
        cell = ws_wbs.cell(row=row, column=1, value=item['id'])
        cell.font = font
        cell.border = border
        if fill:
            cell.fill = fill
        
        # Name column with indentation based on level
        indent = "  " * item['level']
        cell = ws_wbs.cell(row=row, column=2, value=f"{indent}{item['name']}")
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True)
        if fill:
            cell.fill = fill

        requirement_excerpt = build_requirement_excerpt(item, requirement_lookup, children_by_parent)
        cell = ws_wbs.cell(row=row, column=3, value=requirement_excerpt)
        cell.font = task_font if item['type'] == 'task' else font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        if fill:
            cell.fill = fill
        
        # Parallel indicator column
        can_parallel = item.get('can_start_parallel', False)
        parallel_text = "🔄" if can_parallel else ""
        cell = ws_wbs.cell(row=row, column=4, value=parallel_text)
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        if fill:
            cell.fill = fill

        # Build the cost formula that references rates from Sheet 2
        cost_formula_parts = []
        is_leaf = item['id'] in leaf_ids

        if is_leaf:
            hours_by_role = distribute_hours_by_role(item['hours'], item['skills'])
        else:
            hours_by_role = {}

        # Role hours columns (start from column 5 now)
        for col, role in enumerate(all_roles, 5):
            col_letter = get_column_letter(col)

            if is_leaf:
                hours = hours_by_role.get(role, 0)
                cell_value = round(hours, 1) if hours > 0 else 0
            else:
                child_rows = [row_by_item_id[child['id']] for child in children_by_parent.get(item['id'], [])]
                cell_value = build_sum_formula(col_letter, child_rows)
                hours = None

            cell = ws_wbs.cell(row=row, column=col, value=cell_value)
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if fill:
                cell.fill = fill
            # Hide zero values on leaf rows.
            if is_leaf and hours == 0:
                cell.number_format = '0.0;0.0;'  # Format to hide zero values

            # Build formula part: hours * VLOOKUP for rate
            if is_leaf and hours > 0:
                # VLOOKUP formula to find rate for this role in 'Ставки профессий' sheet
                formula_part = f"IF({col_letter}{row}=0,0,{col_letter}{row}*VLOOKUP(\"{role}\",'Ставки профессий'!$A:$B,2,FALSE))"
                cost_formula_parts.append(formula_part)
        
        # Cost column with formula
        cost_col = len(headers)
        if is_leaf and cost_formula_parts:
            cost_formula = "=" + "+".join(cost_formula_parts)
            cell = ws_wbs.cell(row=row, column=cost_col, value=cost_formula)
        elif is_leaf:
            cell = ws_wbs.cell(row=row, column=cost_col, value=0)
        else:
            child_rows = [row_by_item_id[child['id']] for child in children_by_parent.get(item['id'], [])]
            cost_formula = build_sum_formula(get_column_letter(cost_col), child_rows)
            cell = ws_wbs.cell(row=row, column=cost_col, value=cost_formula)
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal='right')
        cell.number_format = '#,##0'
        if fill:
            cell.fill = fill
        
        row += 1
    
    data_end_row = row - 1
    
    # Add totals row
    row += 1
    totals_row = row
    
    # Label
    cell = ws_wbs.cell(row=row, column=1, value="")
    cell.fill = total_fill
    cell.border = border

    cell = ws_wbs.cell(row=row, column=2, value="ИТОГО")
    cell.font = total_font
    cell.fill = total_fill
    cell.border = border

    # Parallel column in totals
    cell = ws_wbs.cell(row=row, column=3, value="")
    cell.fill = total_fill
    cell.border = border

    cell = ws_wbs.cell(row=row, column=4, value="")
    cell.fill = total_fill
    cell.border = border

    leaf_rows = [row_by_item_id[item_id] for item_id in sorted(leaf_ids, key=lambda item_id: row_by_item_id[item_id])]

    # Role totals with SUM formulas (start from column 5)
    for col, role in enumerate(all_roles, 5):
        col_letter = get_column_letter(col)
        formula = build_sum_formula(col_letter, leaf_rows)
        cell = ws_wbs.cell(row=row, column=col, value=formula)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = '#,##0.0'
    
    # Total cost with SUM formula
    cost_col_letter = get_column_letter(len(headers))
    cost_formula = build_sum_formula(cost_col_letter, leaf_rows)
    cell = ws_wbs.cell(row=row, column=len(headers), value=cost_formula)
    cell.font = total_font
    cell.fill = total_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='right')
    cell.number_format = '#,##0'
    
    ws_wbs.freeze_panes = 'E2'
    ws_wbs.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{data_end_row}"
    
    # === Sheet 2: Role Rates ===
    ws_rates = wb.create_sheet(title="Ставки профессий")
    
    # Headers
    headers_rates = ["Профессия (роль)", "Ставка в час (руб)"]
    for col, header in enumerate(headers_rates, 1):
        cell = ws_rates.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws_rates.column_dimensions['A'].width = 30
    ws_rates.column_dimensions['B'].width = 20
    
    # Add only the roles that are actually used in this WBS.
    rate_row = 2

    for role in all_roles:
        cell = ws_rates.cell(row=rate_row, column=1, value=role)
        cell.border = border
        
        rate = get_role_rate(role, rates)
        cell = ws_rates.cell(row=rate_row, column=2, value=rate)
        cell.border = border
        cell.alignment = Alignment(horizontal='right')
        cell.number_format = '#,##0'
        
        rate_row += 1
    
    # Add note about editing rates
    rate_row += 2
    cell = ws_rates.cell(row=rate_row, column=1, value="Примечание: Измените ставки на этом листе для автоматического пересчета стоимости на листе WBS.")
    cell.font = Font(italic=True, color="666666")
    ws_rates.merge_cells(start_row=rate_row, start_column=1, end_row=rate_row, end_column=2)
    
    # Freeze header row
    ws_rates.freeze_panes = 'A2'
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    logger.info(f"WBS Excel export created with {len(work_items)} work items and {len(all_roles)} roles")
    
    return output


def _prepare_gantt_data(wbs_data: dict) -> list:
    """Prepare data for Gantt chart visualization.
    
    Calculates start days based on dependencies and parallel execution capability.
    Tasks marked as can_start_parallel will start at the same time as other parallel tasks.
    
    Args:
        wbs_data: WBS data dictionary
        
    Returns:
        List of items with start_day, duration_days, and dependencies
    """
    items = []
    
    if not wbs_data or 'phases' not in wbs_data:
        return items
    
    item_end_days = {}
    current_day = 0

    for phase in wbs_data.get('phases', []):
        phase_id = phase.get('id', '')
        phase_duration = phase.get('duration', 0) or phase.get('duration_days', 0)
        if not phase_duration:
            hours = phase.get('estimated_hours', 0)
            phase_duration = max(1, hours // 8)

        phase_start = current_day
        parallel_wp_end = phase_start
        sequential_wp_start = phase_start
        phase_items = []

        for wp in phase.get('work_packages', []):
            wp_id = wp.get('id', '')
            wp_duration = wp.get('duration_days', 0)
            if not wp_duration:
                hours = wp.get('estimated_hours', 0)
                wp_duration = max(1, hours // 8)

            wp_dependencies = wp.get('dependencies', [])
            can_parallel = wp.get('can_start_parallel', False)

            dep_end_day = 0
            for dep_id in wp_dependencies:
                if dep_id in item_end_days:
                    dep_end_day = max(dep_end_day, item_end_days[dep_id])

            if can_parallel:
                wp_start = dep_end_day if dep_end_day > 0 else phase_start
            else:
                wp_start = max(sequential_wp_start, dep_end_day)

            task_items = []
            parallel_task_end = wp_start
            sequential_task_start = wp_start

            for task in wp.get('tasks', []):
                task_id = task.get('id', '')
                task_duration = task.get('duration_days', 0)
                if not task_duration:
                    hours = task.get('estimated_hours', 0)
                    task_duration = max(1, hours // 8)

                task_dependencies = task.get('dependencies', [])
                task_can_parallel = task.get('can_start_parallel', False)

                task_dep_end = 0
                for dep_id in task_dependencies:
                    if dep_id in item_end_days:
                        task_dep_end = max(task_dep_end, item_end_days[dep_id])

                if task_can_parallel:
                    task_start = task_dep_end if task_dep_end > 0 else wp_start
                else:
                    task_start = max(sequential_task_start, task_dep_end)

                task_end = task_start + task_duration
                item_end_days[task_id] = task_end

                if not task_can_parallel:
                    sequential_task_start = task_end

                parallel_task_end = max(parallel_task_end, task_end)
                task_items.append({
                    'id': task_id,
                    'name': task.get('name', ''),
                    'type': 'task',
                    'level': 2,
                    'start_day': task_start,
                    'duration_days': task_duration,
                    'dependencies': task_dependencies,
                    'can_parallel': task_can_parallel
                })

            wp_end = max(wp_start + wp_duration, parallel_task_end)
            item_end_days[wp_id] = wp_end

            if not can_parallel:
                sequential_wp_start = wp_end

            parallel_wp_end = max(parallel_wp_end, wp_end)
            phase_items.append({
                'work_package': {
                    'id': wp_id,
                    'name': wp.get('name', ''),
                    'type': 'work_package',
                    'level': 1,
                    'start_day': wp_start,
                    'duration_days': max(1, wp_end - wp_start),
                    'dependencies': wp_dependencies,
                    'can_parallel': can_parallel
                },
                'tasks': task_items
            })

        actual_phase_end = parallel_wp_end if phase.get('work_packages') else phase_start + phase_duration
        item_end_days[phase_id] = actual_phase_end
        current_day = actual_phase_end

        items.append({
            'id': phase_id,
            'name': phase.get('name', ''),
            'type': 'phase',
            'level': 0,
            'start_day': phase_start,
            'duration_days': max(1, actual_phase_end - phase_start),
            'dependencies': [],
            'can_parallel': False
        })

        for wp_entry in phase_items:
            items.append(wp_entry['work_package'])
            items.extend(wp_entry['tasks'])

    return items


def export_wbs_to_excel(result_data: dict, custom_rates: dict = None) -> tuple:
    """Export WBS to Excel file.
    
    Args:
        result_data: Analysis result data
        custom_rates: Optional dictionary of custom role rates
        
    Returns:
        Tuple of (BytesIO object, filename)
    """
    excel_file = create_wbs_excel(result_data, custom_rates)
    filename = "wbs_export.xlsx"
    
    return excel_file, filename
